from openpyxl import Workbook, load_workbook
from openpyxl.utils import  get_column_letter
from faker import Faker

# Variables
names = []
Fake = Faker()

# Generate Fake Names
for i in range(101):
    name = Fake.name().split(' ')
    names.append(name)

# Create, Update and Save workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = 'Fake Names'

for name_tuple in names:
    ws.append(name_tuple)

wb.save('FakeData.xlsx')


wb = load_workbook('FakeData.xlsx')
ws = wb.active

names = []

for row in range (1,50):
    name_l = []
    for col in range (1,3):
        char = get_column_letter(col)
        name_l.append(ws[char+str(row)].value)
    names.append(name_l)








